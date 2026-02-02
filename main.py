"""
FedEx Tracker API - BloomsPal SonIA
Backend FastAPI Application
"""

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
import tempfile
import uuid

# CONFIGURACION - CREDENCIALES FEDEX
FEDEX_API_KEY = os.getenv("FEDEX_API_KEY", "l7e4ca666923294740bae8dfde52ca1f52")
FEDEX_SECRET_KEY = os.getenv("FEDEX_SECRET_KEY", "81d7f9db60554e9b97ffa7c76075763c")
FEDEX_ACCOUNT = os.getenv("FEDEX_ACCOUNT", "740561073")
FEDEX_BASE_URL = "https://apis.fedex.com"

app = FastAPI(
    title="FedEx Tracker - BloomsPal SonIA",
    description="API para rastrear envios FedEx y generar reportes Excel",
    version="1.0.0"
)

templates = Jinja2Templates(directory="templates")
TEMP_DIR = tempfile.gettempdir()


class FedExClient:
    """Cliente para FedEx Track API"""

    def __init__(self):
        self.access_token = None

    def authenticate(self):
        """Obtener token OAuth 2.0"""
        url = f"{FEDEX_BASE_URL}/oauth/token"
        response = requests.post(
            url,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={
                "grant_type": "client_credentials",
                "client_id": FEDEX_API_KEY,
                "client_secret": FEDEX_SECRET_KEY
            },
            timeout=30
        )
        if response.status_code == 200:
            self.access_token = response.json().get("access_token")
            return True
        return False

    def track_multiple(self, tracking_numbers):
        """Rastrear multiples envios"""
        if not self.access_token:
            if not self.authenticate():
                raise HTTPException(status_code=401, detail="Error de autenticacion con FedEx")
        url = f"{FEDEX_BASE_URL}/track/v1/trackingnumbers"
        tracking_info = [
            {"trackingNumberInfo": {"trackingNumber": tn}}
            for tn in tracking_numbers[:30]
        ]
        response = requests.post(
            url,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.access_token}",
                "X-locale": "en_US"
            },
            json={
                "includeDetailedScans": True,
                "trackingInfo": tracking_info
            },
            timeout=60
        )
        if response.status_code == 200:
            return response.json()
        return None

    def parse_results(self, result):
        """Parsear resultados del API"""
        parsed = []
        if not result or "output" not in result:
            return parsed
        for track_result in result.get("output", {}).get("completeTrackResults", []):
            tracking_number = track_result.get("trackingNumber", "")
            for detail in track_result.get("trackResults", []):
                latest = detail.get("latestStatusDetail", {})
                status_code = latest.get("code", "")
                status_desc = latest.get("description", "")
                status_map = {
                    "DL": "Delivered", "IT": "In Transit", "PU": "Picked Up",
                    "OD": "Out for Delivery", "DE": "Delivery Exception",
                    "SE": "Shipment Exception", "CD": "Clearance Delay", "OC": "Label Created"
                }
                status = status_map.get(status_code, status_desc or status_code)
                dates = {dt.get("type"): dt.get("dateTime") for dt in detail.get("dateAndTimes", [])}
                ship_date = dates.get("SHIP") or dates.get("ACTUAL_PICKUP")
                delivery_date = dates.get("ACTUAL_DELIVERY")
                label_date = None
                for event in detail.get("scanEvents", []):
                    if "information sent" in event.get("eventDescription", "").lower():
                        label_date = event.get("date")
                        break
                origin = detail.get("originLocation", {}).get("locationContactAndAddress", {}).get("address", {})
                parsed.append({
                    "tracking": tracking_number,
                    "status": status,
                    "label_date": label_date,
                    "ship_date": ship_date,
                    "delivery_date": delivery_date,
                    "origin_city": origin.get("city", ""),
                    "origin_state": origin.get("stateOrProvinceCode", ""),
                    "origin_country": origin.get("countryCode", "")
                })
        return parsed


def working_days(start, end):
    """Calcular dias habiles"""
    if not start or not end:
        return 0
    days = 0
    current = start + timedelta(days=1)
    while current <= end:
        if current.weekday() < 5:
            days += 1
        current += timedelta(days=1)
    return days


def parse_date(date_str):
    """Convertir string a datetime"""
    if not date_str:
        return None
    try:
        if 'T' in str(date_str):
            return datetime.fromisoformat(str(date_str).replace('Z', '+00:00').split('+')[0])
        parts = str(date_str).split('/')
        if len(parts) == 3:
            m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
            return datetime(y + 2000 if y < 100 else y, m, d)
    except Exception:
        pass
    return None


def generate_excel(data, client_map):
    """Generar reporte Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "FedEx Tracking Report"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    delivered_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    label_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    exception_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    transit_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    headers = [
        "Cliente", "Tracking", "Status", "Label Date", "Ship Date",
        "Days After Ship", "Working Days", "Days After Label",
        "Origin City", "Origin State", "Origin Country", "SonIA"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    widths = [18, 16, 16, 12, 12, 14, 14, 20, 12, 10, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    today = datetime.now()
    for row_idx, item in enumerate(data, 2):
        tn = item['tracking']
        status = item['status']
        label_dt = parse_date(item.get('label_date'))
        ship_dt = parse_date(item.get('ship_date'))
        deliv_dt = parse_date(item.get('delivery_date'))
        if "Delivered" in status and deliv_dt and ship_dt:
            days_ship = (deliv_dt - ship_dt).days
            work_days = working_days(ship_dt, deliv_dt)
            days_label = (deliv_dt - label_dt).days if label_dt else 0
            label_text = f"DELIVERED in {days_label} days"
            sonia = f"Entregado en {days_label} dias"
        elif "Label" in status or "Created" in status:
            days_ship = 0
            work_days = 0
            days_label = (today - label_dt).days if label_dt else 0
            label_text = str(days_label)
            sonia = f"Pendiente - {days_label} dias"
        elif "Transit" in status:
            days_ship = (today - ship_dt).days if ship_dt else 0
            work_days = working_days(ship_dt, today) if ship_dt else 0
            days_label = (today - label_dt).days if label_dt else 0
            label_text = str(days_label)
            sonia = f"En transito - {days_ship} dias"
        elif "Exception" in status or "Delay" in status:
            days_ship = (today - ship_dt).days if ship_dt else 0
            work_days = working_days(ship_dt, today) if ship_dt else 0
            days_label = (today - label_dt).days if label_dt else 0
            label_text = str(days_label)
            sonia = "EXCEPCION - REVISAR"
        else:
            days_ship, work_days, days_label = 0, 0, 0
            label_text = "0"
            sonia = "Verificar"
        row_data = [
            client_map.get(tn, ""),
            tn,
            status,
            label_dt.strftime('%Y-%m-%d') if label_dt else '',
            ship_dt.strftime('%Y-%m-%d') if ship_dt else '',
            days_ship,
            work_days,
            label_text,
            item.get('origin_city', ''),
            item.get('origin_state', ''),
            item.get('origin_country', ''),
            sonia
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if "Delivered" in status:
                cell.fill = delivered_fill
            elif "Label" in status or "Created" in status:
                cell.fill = label_fill
            elif "Exception" in status:
                cell.fill = exception_fill
            elif "Transit" in status:
                cell.fill = transit_fill
    ws.freeze_panes = 'A2'
    filename = f"FedEx_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(TEMP_DIR, filename)
    wb.save(filepath)
    return filepath


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Pagina principal"""
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/api/track")
async def track_shipments(file: UploadFile = File(...)):
    """Procesar archivo Excel y rastrear envios"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")
    temp_input = os.path.join(TEMP_DIR, f"input_{uuid.uuid4().hex}.xlsx")
    try:
        with open(temp_input, "wb") as f:
            content = await file.read()
            f.write(content)
        df = pd.read_excel(temp_input, dtype=str)
        tracking_col = df.columns[14] if len(df.columns) > 14 else df.columns[-1]
        client_col = df.columns[2] if len(df.columns) > 2 else df.columns[0]
        trackings = []
        client_map = {}
        for _, row in df.iterrows():
            tracking = str(row[tracking_col]).strip() if pd.notna(row[tracking_col]) else ''
            client = str(row[client_col]).strip() if pd.notna(row[client_col]) else ''
            if tracking and tracking != 'nan' and len(tracking) >= 10:
                tracking = tracking.replace(' ', '')
                if tracking.isdigit():
                    trackings.append(tracking)
                    client_map[tracking] = client
        if not trackings:
            raise HTTPException(status_code=400, detail="No se encontraron tracking numbers validos")
        fedex = FedExClient()
        all_results = []
        for i in range(0, len(trackings), 30):
            batch = trackings[i:i+30]
            result = fedex.track_multiple(batch)
            if result:
                parsed = fedex.parse_results(result)
                all_results.extend(parsed)
        output_path = generate_excel(all_results, client_map)
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="FedEx_Tracking_Report_SonIA.xlsx"
        )
    finally:
        if os.path.exists(temp_input):
            os.remove(temp_input)


@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "FedEx Tracker - BloomsPal SonIA"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 8000)))
